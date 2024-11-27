import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

planilha_cliente = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_cliente = planilha_cliente['Sheet1']

drive = webdriver.Chrome()
drive.get('https://consultcpf-devaprender.netlify.app/')
sleep(5)

for linha in pagina_cliente.iter_rows(min_row=2,values_only=True):
    nome, valor, cpf, vencimento = linha

    campo_pesquisa = drive.find_element(By.XPATH,'//*[@id="cpfInput"]')
    sleep(1)
    campo_pesquisa.clear()
    campo_pesquisa.send_keys(cpf)
    sleep(1)

    botao_pesquisa = drive.find_element(By.XPATH,'//*[@id="consultaForm"]/button')
    sleep(1)
    botao_pesquisa.click()
    sleep(4)

    status = drive.find_element(By.XPATH,'//*[@id="statusLabel"]')
    sleep(1)
    if status.text == "em dia": 

        data_pagamento = drive.find_element(By.XPATH,'//*[@id="paymentDate"]')
        data_pagamento_limpo = data_pagamento.text.split()[3]

        data_metodo = drive.find_element(By.XPATH,'//*[@id="paymentMethod"]')
        data_metodo_limpo = data_pagamento.text.split()[3]

        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_cliente['Sheet1']

        pagina_fechamento.append([nome, valor, cpf, vencimento, 'em dia',data_pagamento_limpo, data_metodo_limpo ])

        planilha_fechamento.save('planilha fechamento.xlsx')
    else:
        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_cliente['Sheet1']

        pagina_fechamento.append([nome, valor, cpf, vencimento, 'pendente'])

        planilha_fechamento.save('planilha fechamento.xlsx')